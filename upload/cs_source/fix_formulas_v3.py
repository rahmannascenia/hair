#!/usr/bin/env python3
"""
fix_formulas_v3.py — Final fix for ClosedXML formula + chart issues.

PROBLEM:
  ClosedXML 0.104 stores formulas as SHARED STRINGS, not as real formula cells.
  A cell like <c r="J18" t="s"><v>705</v></c> where shared string 705 = "=E18*RATE_A+..."
  is seen by Excel as TEXT, not as a formula. User must click into each cell and press
  Enter to make it compute.

  Additionally, ClosedXML uses the "x:" namespace prefix in all XML elements.

FIX (3 steps):
  1. Walk every worksheet XML. For each <c t="s"><v>N</v></c> cell where shared
     string N starts with "=", rewrite as a real <c><f>FORMULA</f></c> cell.
  2. Evaluate all formulas with the `formulas` Python library to compute values.
  3. Inject cached <v>VALUE</v> next to each <f>.
  4. Set fullCalcOnLoad="1" on workbook.xml.
"""

import os
import re
import sys
import shutil
import zipfile
import html as html_mod
from pathlib import Path

FILE_PATH = '/home/z/my-project/download/BarendraInternational_ERP.xlsx'


def xml_escape(s):
    if not isinstance(s, str):
        return str(s)
    return s.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;').replace('"', '&quot;')


def make_value_xml(value):
    if isinstance(value, bool):
        return f'<v>{1 if value else 0}</v>', ' t="b"'
    if isinstance(value, (int, float)):
        if isinstance(value, float):
            if value != value or value in (float('inf'), float('-inf')):
                return '<v>0</v>', ''
        return f'<v>{value}</v>', ''
    if isinstance(value, str):
        return f'<v>{xml_escape(value)}</v>', ' t="str"'
    return f'<v>{xml_escape(str(value))}</v>', ' t="str"'


def unzip_xlsx(file_path, dest_dir):
    if os.path.exists(dest_dir):
        shutil.rmtree(dest_dir)
    os.makedirs(dest_dir)
    with zipfile.ZipFile(file_path, 'r') as z:
        z.extractall(dest_dir)


def rezip_xlsx(src_dir, file_path):
    tmp_zip = file_path + '.tmp'
    if os.path.exists(tmp_zip):
        os.remove(tmp_zip)
    content_types = os.path.join(src_dir, '[Content_Types].xml')
    with zipfile.ZipFile(tmp_zip, 'w', zipfile.ZIP_DEFLATED) as zout:
        if os.path.exists(content_types):
            zout.write(content_types, '[Content_Types].xml')
        for root, dirs, files in os.walk(src_dir):
            for fname in files:
                full_path = os.path.join(root, fname)
                arcname = os.path.relpath(full_path, src_dir)
                if arcname == '[Content_Types].xml':
                    continue
                zout.write(full_path, arcname)
    shutil.move(tmp_zip, file_path)


def load_shared_strings(tmp_dir):
    ss_path = os.path.join(tmp_dir, 'xl', 'sharedStrings.xml')
    if not os.path.exists(ss_path):
        return []
    with open(ss_path, 'r', encoding='utf-8') as f:
        ss_xml = f.read()
    # Match <si> or <x:si> elements
    sis = re.findall(r'<(?:x:)?si[^>]*>(.*?)</(?:x:)?si>', ss_xml, re.DOTALL)
    result = []
    for si in sis:
        text_m = re.search(r'<(?:x:)?t[^>]*>(.*?)</(?:x:)?t>', si, re.DOTALL)
        if text_m:
            result.append(html_mod.unescape(text_m.group(1)))
        else:
            result.append('')
    return result


def get_sheet_file_mapping(tmp_dir):
    wb_xml_path = os.path.join(tmp_dir, 'xl', 'workbook.xml')
    with open(wb_xml_path, 'r', encoding='utf-8') as f:
        wb_xml = f.read()

    # Handle x: namespace prefix on sheet elements
    sheets = re.findall(r'<(?:x:)?sheet[^>]*name="([^"]+)"[^>]*sheetId="(\d+)"[^>]*r:id="([^"]+)"', wb_xml)
    if not sheets:
        sheets = re.findall(r'<(?:x:)?sheet[^>]*name="([^"]+)"[^>]*r:id="([^"]+)"[^>]*sheetId="(\d+)"', wb_xml)
        sheets = [(n, sid, rid) for n, rid, sid in sheets]
    sheets = [(html_mod.unescape(n), sid, rid) for n, sid, rid in sheets]

    rels_path = os.path.join(tmp_dir, 'xl', '_rels', 'workbook.xml.rels')
    if not os.path.exists(rels_path):
        rels_path = os.path.join(tmp_dir, 'xl', 'workbook.xml.rels')
    with open(rels_path, 'r', encoding='utf-8') as f:
        rels_xml = f.read()
    rid_to_target = dict(re.findall(r'<(?:x:)?Relationship[^>]*Id="([^"]+)"[^>]*Target="([^"]+)"', rels_xml))
    if not rid_to_target:
        for m in re.finditer(r'<(?:x:)?Relationship[^>]*Target="([^"]+)"[^>]*Id="([^"]+)"', rels_xml):
            target, rid = m.groups()
            rid_to_target[rid] = target

    result = []
    for sheet_name, sheet_id, rid in sheets:
        if rid not in rid_to_target:
            continue
        target = rid_to_target[rid]
        if target.startswith('/'):
            target = target.lstrip('/')
        sheet_xml_path = os.path.join(tmp_dir, target)
        if not os.path.exists(sheet_xml_path):
            sheet_xml_path = os.path.join(tmp_dir, 'xl', 'worksheets', os.path.basename(target))
        if os.path.exists(sheet_xml_path):
            result.append((sheet_name, sheet_xml_path))
    return result


# ──────────────────────────────────────────────────────────────────────────
#  STEP 1: Convert shared-string formulas to real <f> formula cells
# ──────────────────────────────────────────────────────────────────────────
def convert_formulas(sheet_xml, shared_strings):
    """Convert cells that reference shared strings containing formula text (starting with '=')
    into real formula cells. Handles x: namespace prefix."""

    count = 0

    # Pattern: <x:c r="CELL" ... t="s"><x:v>INDEX</x:v></x:c>
    # or <c r="CELL" ... t="s"><v>INDEX</v></c>
    cell_pattern = re.compile(
        r'<(?P<ns>x:)?c\b([^>]*\br="(?P<cell>[A-Z]+\d+)"[^>]*\bt="s"[^>]*)>'
        r'\s*<(?P=ns)?v[^>]*>(?P<idx>\d+)</(?P=ns)?v>\s*'
        r'</(?P=ns)?c>',
        re.DOTALL,
    )

    def replace_shared(m):
        nonlocal count
        ns = m.group('ns') or ''
        attrs = m.group(2)
        cell_ref = m.group('cell')
        str_idx = int(m.group('idx'))
        if str_idx >= len(shared_strings):
            return m.group(0)
        text = shared_strings[str_idx]
        if not text.startswith('='):
            return m.group(0)
        formula_text = text[1:]
        count += 1
        s_match = re.search(r'\bs="(\d+)"', attrs)
        s_attr = f' s="{s_match.group(1)}"' if s_match else ''
        return f'<{ns}c r="{cell_ref}"{s_attr}><{ns}f>{xml_escape(formula_text)}</{ns}f></{ns}c>'

    new_xml = cell_pattern.sub(replace_shared, sheet_xml)
    return new_xml, count


def step1_convert_formulas(file_path):
    print('\nStep 1: Convert shared-string formulas to real <f> formula cells...')
    tmp_dir = '/tmp/xlsx_step1'
    unzip_xlsx(file_path, tmp_dir)

    shared_strings = load_shared_strings(tmp_dir)
    print(f'  Loaded {len(shared_strings)} shared strings')

    sheet_mapping = get_sheet_file_mapping(tmp_dir)
    total = 0
    for sheet_name, sheet_xml_path in sheet_mapping:
        with open(sheet_xml_path, 'r', encoding='utf-8') as f:
            sheet_xml = f.read()
        new_xml, count = convert_formulas(sheet_xml, shared_strings)
        if count > 0:
            with open(sheet_xml_path, 'w', encoding='utf-8') as f:
                f.write(new_xml)
            print(f'  {sheet_name}: {count} formulas converted')
            total += count
    print(f'  Total: {total} formulas converted')
    rezip_xlsx(tmp_dir, file_path)
    return total


# ──────────────────────────────────────────────────────────────────────────
#  STEP 2: Evaluate all formulas
# ──────────────────────────────────────────────────────────────────────────
def step2_evaluate(file_path):
    print('\nStep 2: Evaluate all formulas with `formulas` library...')
    try:
        import formulas
    except ImportError:
        print('  Installing formulas library...')
        import subprocess
        subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'formulas'])
        import formulas

    abs_path = os.path.abspath(file_path)
    xl_model = formulas.ExcelModel().loads(abs_path).finish()
    print('  Solving...')
    solution = xl_model.calculate()
    print(f'  Evaluation complete. {len(solution)} cells in solution.')

    file_basename = os.path.basename(abs_path).upper()
    result = {}
    key_pattern = re.compile(r"'\[([^\]]+)\]([^']+)'!([A-Z]+\d+)")
    import numpy as np
    for key, val in solution.items():
        if not isinstance(key, str):
            continue
        m = key_pattern.match(key)
        if not m:
            continue
        fname, sheet, cell = m.groups()
        if fname.upper() != file_basename:
            continue
        try:
            if hasattr(val, 'value'):
                v = val.value
            else:
                v = val
            if isinstance(v, np.ndarray):
                if v.size == 1:
                    v = v.item()
                elif v.size == 0:
                    v = None
                else:
                    v = v.flatten()[0]
            if isinstance(v, str) and (v.startswith('#') or v.startswith('=')):
                v = None
            if isinstance(v, (np.integer,)):
                v = int(v)
            elif isinstance(v, (np.floating,)):
                v = float(v)
            elif isinstance(v, (np.bool_,)):
                v = bool(v)
        except Exception:
            v = None
        if v is not None:
            result.setdefault(sheet.upper(), {})[cell] = v
    total = sum(len(v) for v in result.values())
    print(f'  Computed values for {total} cells across {len(result)} sheets')
    return result


# ──────────────────────────────────────────────────────────────────────────
#  STEP 3: Inject cached values
# ──────────────────────────────────────────────────────────────────────────
def inject_cached_values(sheet_xml, computed_for_sheet):
    count = 0
    skipped = 0

    # Pattern: <x:c r="CELL" [s="N"]><x:f>FORMULA</x:f></x:c>
    cell_pattern = re.compile(
        r'(<(?P<ns>x:)?c\b[^>]*\br="(?P<cell>[A-Z]+\d+)"[^>]*>)<(?P=ns)f>([^<]*)</(?P=ns)f>(</(?P=ns)c>)',
    )

    def replace_cell(m):
        nonlocal count, skipped
        ns = m.group('ns') or ''
        opening_tag = m.group(1)
        cell_ref = m.group('cell')
        formula_text = m.group(4)
        closing_tag = m.group(5)

        value = computed_for_sheet.get(cell_ref)
        if value is None:
            skipped += 1
            return m.group(0)

        value_xml, type_attr = make_value_xml(value)
        if type_attr and ' t="' not in opening_tag:
            new_opening = opening_tag[:-1] + type_attr + '>'
        else:
            new_opening = opening_tag
        # Add namespace prefix to <v>
        ns_v = value_xml.replace('<v>', f'<{ns}v>').replace('</v>', f'</{ns}v>')
        count += 1
        return f'{new_opening}<{ns}f>{formula_text}</{ns}f>{ns_v}{closing_tag}'

    new_xml = cell_pattern.sub(replace_cell, sheet_xml)
    return new_xml, count, skipped


def step3_inject(file_path, computed_values):
    print('\nStep 3: Inject cached values into real formula cells...')
    tmp_dir = '/tmp/xlsx_step3'
    unzip_xlsx(file_path, tmp_dir)

    sheet_mapping = get_sheet_file_mapping(tmp_dir)
    total_injected = 0
    total_skipped = 0
    for sheet_name, sheet_xml_path in sheet_mapping:
        with open(sheet_xml_path, 'r', encoding='utf-8') as f:
            sheet_xml = f.read()
        computed = computed_values.get(sheet_name.upper(), {})
        new_xml, injected, skipped = inject_cached_values(sheet_xml, computed)
        if injected > 0 or skipped > 0:
            with open(sheet_xml_path, 'w', encoding='utf-8') as f:
                f.write(new_xml)
            print(f'  {sheet_name}: {injected} injected, {skipped} skipped')
            total_injected += injected
            total_skipped += skipped
    print(f'  Total: {total_injected} injected, {total_skipped} skipped')

    # Set fullCalcOnLoad on workbook.xml (handle x: namespace)
    wb_xml_path = os.path.join(tmp_dir, 'xl', 'workbook.xml')
    with open(wb_xml_path, 'r', encoding='utf-8') as f:
        wb_xml = f.read()
    # Find existing calcPr (with or without x: prefix)
    calcpr_pattern = re.compile(r'<(?:x:)?calcPr\b[^>]*/>')
    if calcpr_pattern.search(wb_xml):
        m = calcpr_pattern.search(wb_xml)
        existing = m.group(0)
        if 'fullCalcOnLoad' in existing:
            new = re.sub(r'fullCalcOnLoad="[^"]*"', 'fullCalcOnLoad="1"', existing)
        else:
            new = existing[:-2] + ' fullCalcOnLoad="1"/>'
        wb_xml = wb_xml.replace(existing, new)
    else:
        # Insert before </x:workbook> or </workbook>
        for close_tag in ['</x:workbook>', '</workbook>']:
            if close_tag in wb_xml:
                wb_xml = wb_xml.replace(close_tag, '<x:calcPr calcId="0" fullCalcOnLoad="1"/>' + close_tag)
                break
    with open(wb_xml_path, 'w', encoding='utf-8') as f:
        f.write(wb_xml)
    print('  Set fullCalcOnLoad="1"')

    rezip_xlsx(tmp_dir, file_path)
    return total_injected


# ──────────────────────────────────────────────────────────────────────────
#  VERIFICATION
# ──────────────────────────────────────────────────────────────────────────
def verify(file_path):
    from openpyxl import load_workbook
    wb = load_workbook(file_path, data_only=True)
    wb_f = load_workbook(file_path, data_only=False)
    checks = [
        ('Procurement & LC', 'I6', 'Goods USD'),
        ('Procurement & LC', 'N6', 'Landed BDT'),
        ('F-01 Fatema', 'J18', 'Base wage (worker 1)'),
        ('F-01 Fatema', 'K18', 'Attendance bonus'),
        ('F-01 Fatema', 'L18', 'Total payable'),
        ('F-01 Fatema', 'C11', 'Supervisor pay'),
        ('Payroll Summary', 'K5', 'F-01 workers total'),
        ('Payroll Summary', 'M5', 'F-01 grand total'),
        ('Dashboard', 'B6', 'Total Lots tile'),
        ('Lot Master', 'H5', 'Total Raw Cost'),
    ]
    print('\nVerification:')
    print(f'  {"Cell":45s} | {"Cached Value":20s}')
    print(f'  {"-"*45} | {"-"*20}')
    ok = 0
    for sheet, cell, label in checks:
        if sheet not in wb.sheetnames:
            print(f'  SKIP {sheet}!{cell} — sheet not found')
            continue
        cached = wb[sheet][cell].value
        formula = wb_f[sheet][cell].value
        v_str = str(cached)[:18] if cached is not None else 'BLANK!'
        marker = 'OK' if cached is not None else 'FAIL'
        if cached is not None:
            ok += 1
        print(f'  {marker} {sheet+"!"+cell+" ("+label+")":45s} | {v_str:20s}')
    print(f'\n  {ok}/{len(checks)} cells have cached values')

    # Count charts
    chart_count = 0
    for sn in wb.sheetnames:
        ws = wb[sn]
        charts = getattr(ws, '_charts', [])
        chart_count += len(charts)
    print(f'  Charts in workbook: {chart_count}')
    return ok == len(checks)


# ──────────────────────────────────────────────────────────────────────────
#  MAIN
# ──────────────────────────────────────────────────────────────────────────
def main():
    print(f'Fixing formulas in: {FILE_PATH}')
    step1_convert_formulas(FILE_PATH)
    computed = step2_evaluate(FILE_PATH)
    step3_inject(FILE_PATH, computed)
    print('\nStep 4: Verification...')
    verify(FILE_PATH)
    print(f'\nFile size: {os.path.getsize(FILE_PATH):,} bytes')
    print('Done!')


if __name__ == '__main__':
    main()
